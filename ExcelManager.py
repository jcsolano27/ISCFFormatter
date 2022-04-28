import openpyxl


class ExcelManager:

    path = ""
    wb = None
    ws = None
    xlsx_name = ""
    current_row = 1
    max_columns_num = 0
    title_columns_num = 0

    def __init__(self, output="output", separator=","):
        self.separator = separator
        self.path = output

    def open_file(self, filename):
        self.xlsx_name = filename.split(".")[0]
        self.wb = openpyxl.Workbook()

    def close_file(self):
        self.add_max_num_columns()
        i_ws = self.wb.get_sheet_by_name("Sheet")
        self.wb.remove_sheet(i_ws)
        self.wb.save(f"{self.path}\\{self.xlsx_name}.xlsx")

    def add_sheet(self, sheet_name):
        if self.ws is not None:
            self.add_max_num_columns()
        self.current_row = 1
        self.ws = self.wb.create_sheet(sheet_name)
        self.ws.title = sheet_name

    def add_max_num_columns(self):
        if self.max_columns_num > self.title_columns_num:
            self.ws.cell(row=1, column=self.title_columns_num+1).value = f"(Bracket max amount of columns:{self.max_columns_num})"
            self.title_columns_num = 1
            self.max_columns_num = 1

    def add_row(self, values):
        if not values:
            return

        values_arr = values.split(self.separator)
        columns_num = len(values_arr)

        if self.current_row == 1:
            self.title_columns_num = columns_num
        if self.max_columns_num < columns_num:
            self.max_columns_num = columns_num

        for current_column in range(1, columns_num+1):
            value = values_arr[current_column-1]
            if value.isdecimal():
                self.ws.cell(row=self.current_row, column=current_column).value =  f"=VALUE({values_arr[current_column - 1]})"
            else:
                self.ws.cell(row=self.current_row, column=current_column).value = values_arr[current_column-1]
        self.current_row += 1
